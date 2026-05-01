import csv
import io
from typing import List, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from backend.models.invoice import InvoiceData
from backend.config import config


class ExportService:
    COLUMNS = [
        ("file_name",           "File Name"),
        ("invoice_number",      "Invoice Nr."),
        ("invoice_date",        "Invoice Date"),
        ("vendor_name",         "Vendor"),
        ("vendor_address",      "Vendor Address"),
        ("vendor_iban",         "IBAN"),
        ("vendor_vat_uid",      "VAT UID"),
        ("debtor_name",         "Recipient"),
        ("debtor_address",      "Recipient Address"),
        ("client_number",       "Client Nr."),
        ("currency",            "Currency"),
        ("subtotal",            "Subtotal"),
        ("vat_rate",            "VAT %"),
        ("vat_amount",          "VAT Amount"),
        ("total",               "Total"),
        ("reference_number",    "QR Reference"),
        ("source_type",         "Source"),
        ("confidence_score",    "Confidence %"),
    ]

    @staticmethod
    def _get_value(inv, field_name):
        """Get field value from either a dict or an object."""
        if isinstance(inv, dict):
            return inv.get(field_name)
        return getattr(inv, field_name, None)

    def to_csv(self, invoices: List) -> io.StringIO:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=config.CSV_DELIMITER, quotechar='"', quoting=csv.QUOTE_MINIMAL)
        headers = [col[1] for col in self.COLUMNS]
        writer.writerow(headers)
        for inv in invoices:
            row = []
            for field_name, _ in self.COLUMNS:
                value = self._get_value(inv, field_name)
                if value is None:
                    value = ""
                elif isinstance(value, float):
                    value = f"{value:.2f}"
                row.append(str(value))
            writer.writerow(row)
        output.seek(0)
        return output

    def to_excel(self, invoices: List) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        for col_idx, (_, display_name) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, inv in enumerate(invoices, 2):
            for col_idx, (field_name, _) in enumerate(self.COLUMNS, 1):
                value = self._get_value(inv, field_name)
                if value is None:
                    value = ""
                elif isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if field_name in ("subtotal", "vat_amount", "total"):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
        for col_idx, (_, display_name) in enumerate(self.COLUMNS, 1):
            max_len = len(display_name)
            for row in range(2, len(invoices) + 2):
                cell_val = str(ws.cell(row=row, column=col_idx).value or "")
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output